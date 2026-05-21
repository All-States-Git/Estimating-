#!/usr/bin/env python3
"""
ASM Preconstruction Pipeline Tracker V9
All States Mechanical — Excel workbook generator
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles.differential import DifferentialStyle
from datetime import date

OUTPUT = r"C:\Dev\ASM_Preconstruction_Pipeline_Tracker_V9.xlsx"
TODAY  = date(2026, 5, 5)

# ─── Palette ──────────────────────────────────────────────────────────────────
NAVY    = "1F3864";  BLUE    = "2E75B6";  LTBLUE  = "D9E2F3"
DKGRN   = "375623";  LTGRN   = "E2EFDA";  DKRED   = "C00000"
LTRED   = "FFE0E0";  ORANGE  = "C55A11";  LTORANG = "FCE4D6"
DKYEL   = "7F6000";  LTYEL   = "FFEB9C";  PURPLE  = "4B0082"
LTPURP  = "EAD5F5";  TEAL    = "1F6B75";  LTTEAL  = "D4EEF1"
WHITE   = "FFFFFF";  LGRAY   = "F2F2F2";  MGRAY   = "D9D9D9"
DGRAY   = "595959";  BLACK   = "000000"

def F(size=10, bold=False, color=BLACK, italic=False):
    return Font(name="Arial", size=size, bold=bold, color=color, italic=italic)

def BG(color): return PatternFill("solid", fgColor=color)

def AL(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def BD(color=MGRAY, style="thin"):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def hcell(cell, text, bg=NAVY, fg=WHITE, size=10, bold=True, wrap=True, align="center"):
    cell.value = text
    cell.font  = Font(name="Arial", size=size, bold=bold, color=fg)
    cell.fill  = BG(bg)
    cell.alignment = AL(align, "center", wrap)
    cell.border = BD()

def vcell(cell, val, bg=None, fg=BLACK, size=10, bold=False, num_fmt=None,
          align="left", wrap=False):
    cell.value = val
    cell.font  = Font(name="Arial", size=size, bold=bold, color=fg)
    if bg: cell.fill = BG(bg)
    cell.alignment = AL(align, "center", wrap)
    cell.border = BD()
    if num_fmt: cell.number_format = num_fmt

def section_bar(ws, row, col_from, col_to, text, bg=NAVY, fg=WHITE, height=22):
    ws.merge_cells(f"{col_from}{row}:{col_to}{row}")
    c = ws[f"{col_from}{row}"]
    c.value = text
    c.font  = Font(name="Arial", size=11, bold=True, color=fg)
    c.fill  = BG(bg)
    c.alignment = AL("left", "center")
    c.border = BD()
    ws.row_dimensions[row].height = height

def title_bar(ws, row, col_from, col_to, text, bg=NAVY, fg=WHITE, size=14, height=32):
    ws.merge_cells(f"{col_from}{row}:{col_to}{row}")
    c = ws[f"{col_from}{row}"]
    c.value = text
    c.font  = Font(name="Arial", size=size, bold=True, color=fg)
    c.fill  = BG(bg)
    c.alignment = AL("center", "center")
    ws.row_dimensions[row].height = height

# ══════════════════════════════════════════════════════════════════════════════
# SETTINGS TAB
# ══════════════════════════════════════════════════════════════════════════════
SETTINGS_LISTS = [
    ("ESTIMATORS", "A", [
        "Mike Johnson", "Sarah Chen", "Tom Rodriguez",
        "Dave Williams", "Lisa Park", "James Carter", "Rachel Nguyen"
    ]),
    ("BID STATUS", "C", [
        "In Progress", "ROM Submitted", "Budget Submitted",
        "Bid Submitted", "Awarded", "Lost", "No Bid", "On Hold"
    ]),
    ("MARKET SEGMENTS", "E", [
        "Multifamily", "Hospitality", "Student Housing",
        "Commercial", "Industrial", "Healthcare",
        "Education", "Design-Build", "Mixed Use", "Civic"
    ]),
    ("SCOPE TYPES", "G", [
        "Plumbing Only", "HVAC Only", "Full Mechanical",
        "Design-Build", "Service", "Budget Only"
    ]),
    ("BID STAGES", "I", [
        "ROM", "Budget", "DD", "Permit Set", "Hard Bid", "GMP", "Design-Build"
    ]),
    ("DELIVERY METHODS", "K", [
        "Hard Bid", "GMP", "Design-Build", "Negotiated", "CM at Risk", "JOC"
    ]),
    ("LOST REASONS", "M", [
        "Price", "Relationship", "Scope Gap", "Schedule Conflict",
        "No Award", "Withdrawn by Owner", "Competition Too Strong",
        "Not Qualified", "Other"
    ]),
    ("GO / NO-GO", "O", ["Go", "Review", "No-Go"]),
    ("GENERAL CONTRACTORS", "Q", [
        "Whiting-Turner", "Turner Construction", "Gilbane Building Co.",
        "Skanska USA", "Balfour Beatty", "DPR Construction",
        "McCarthy Building", "Hensel Phelps", "Brasfield & Gorrie",
        "Clark Construction", "Mortenson Construction", "Barton Malow",
        "Webcor Builders", "JE Dunn Construction", "Swinerton",
        "Suffolk Construction", "Choate Construction",
        "New South Construction", "Owner Direct", "Other"
    ]),
    ("FOLLOW-UP OWNERS", "S", [
        "Mike Johnson", "Sarah Chen", "Tom Rodriguez",
        "Dave Williams", "Lisa Park", "James Carter",
        "Rachel Nguyen", "VP Preconstruction", "President"
    ]),
]

def build_settings(ws):
    ws.sheet_properties.tabColor = "808080"
    title_bar(ws, 1, "A", "V", "ASM PRECONSTRUCTION PIPELINE TRACKER V9 — SETTINGS & DROPDOWN LISTS")
    ws.row_dimensions[2].height = 16
    ws.merge_cells("A2:V2")
    c = ws["A2"]
    c.value = "WARNING: Do not delete, reorder, or rename items in this tab. These lists power all dropdown menus in the Bid Log."
    c.font  = Font(name="Arial", size=10, bold=True, color=ORANGE)
    c.fill  = BG("FFF2CC")
    c.alignment = AL("left", "center")

    for title, col_letter, values in SETTINGS_LISTS:
        ci = column_index_from_string(col_letter)
        ws.row_dimensions[4].height = 22
        hcell(ws.cell(row=4, column=ci), title, size=9, wrap=True)
        for i, val in enumerate(values):
            r = 5 + i
            ws.row_dimensions[r].height = 16
            cell = ws.cell(row=r, column=ci)
            cell.value = val
            cell.font  = F(10)
            cell.fill  = BG(LGRAY if i % 2 == 0 else WHITE)
            cell.alignment = AL("left", "center")
            cell.border = BD()
        ws.column_dimensions[col_letter].width = 22

    for sc in ["B","D","F","H","J","L","N","P","R","T","V"]:
        ws.column_dimensions[sc].width = 1.5

# ══════════════════════════════════════════════════════════════════════════════
# BID LOG TAB
# ══════════════════════════════════════════════════════════════════════════════
BID_COLS = [
    # (display_name,            width,  num_format)
    ("Bid ID",                   9,    "@"),
    ("Project Name",            32,    "@"),
    ("General Contractor",      24,    "@"),
    ("Location",                20,    "@"),
    ("Market Segment",          18,    "@"),
    ("Scope Type",              20,    "@"),
    ("Bid Stage",               14,    "@"),
    ("Delivery Method",         16,    "@"),
    ("Estimator",               16,    "@"),
    ("Bid Received",            14,    "m/d/yyyy"),
    ("Bid Due Date",            14,    "m/d/yyyy"),
    ("Submitted Date",          14,    "m/d/yyyy"),
    ("ASM Bid Amount",          16,    "$#,##0"),
    ("Status",                  16,    "@"),
    ("Awarded Amount",          16,    "$#,##0"),
    ("Winning Competitor",      22,    "@"),
    ("Competitor Amount",       16,    "$#,##0"),
    ("Lost Reason",             18,    "@"),
    ("Expected Award Date",     18,    "m/d/yyyy"),
    ("Probability %",           14,    "0%"),
    ("Next Follow-Up Date",     17,    "m/d/yyyy"),
    ("Follow-Up Owner",         18,    "@"),
    ("Last Contact Date",       16,    "m/d/yyyy"),
    ("Notes",                   38,    "@"),
    ("Go/No-Go Score",          14,    "0"),
    ("Go/No-Go Rec",            14,    "@"),
    ("Days Pending",            14,    "0"),
    ("Follow-Up Status",        16,    "@"),
    ("Data Quality Flag",       44,    "@"),
]

COL_IDX = {name: i+1 for i, (name, _, _) in enumerate(BID_COLS)}

def cl(name): return get_column_letter(COL_IDX[name])

SAMPLE_BIDS = [
    # (bid_id, project, gc, location, market, scope, stage, delivery, estimator,
    #  received, due, submitted, bid_amt, status, award_amt, win_comp, comp_amt,
    #  lost_reason, exp_award, prob, next_fu, fu_owner, last_contact, notes)
    ("ASM-2026-001","The Meridian at Midtown","Whiting-Turner","Atlanta, GA",
     "Multifamily","Full Mechanical","Hard Bid","Hard Bid","Mike Johnson",
     date(2026,3,15),date(2026,4,10),date(2026,4,9),850000,"Bid Submitted",
     None,None,None,None,date(2026,5,20),0.75,
     date(2026,5,12),"Mike Johnson",date(2026,4,28),
     "Strong GC relationship. Two rounds of clarifications received."),

    ("ASM-2026-002","Hampton Inn Downtown Expansion","Turner Construction","Charlotte, NC",
     "Hospitality","Plumbing Only","Permit Set","Hard Bid","Sarah Chen",
     date(2026,2,1),date(2026,3,15),date(2026,3,14),320000,"Awarded",
     320000,None,None,None,1.0,
     None,None,date(2026,4,15),
     "Awarded April 15. Kickoff meeting scheduled for May 20."),

    ("ASM-2026-003","University Commons Phase 2","Gilbane Building Co.","Nashville, TN",
     "Student Housing","Full Mechanical","DD","GMP","Tom Rodriguez",
     date(2026,3,20),date(2026,4,25),date(2026,4,24),1250000,"Bid Submitted",
     None,None,None,None,date(2026,6,1),0.50,
     date(2026,5,15),"Tom Rodriguez",date(2026,5,1),
     "GMP process. Second round of pricing expected in early June."),

    ("ASM-2026-004","Commerce Center Building A","Balfour Beatty","Tampa, FL",
     "Commercial","HVAC Only","Hard Bid","Hard Bid","Dave Williams",
     date(2026,1,10),date(2026,2,20),date(2026,2,19),180000,"Lost",
     None,"Southeastern Mechanical",162000,
     "Price",None,0,
     None,None,date(2026,3,5),
     "Lost on price. Competitor was $18K lower. Review labor pricing."),

    ("ASM-2026-005","The Grove Apartments Phase 3","Mortenson Construction","Denver, CO",
     "Multifamily","Full Mechanical","Hard Bid","Hard Bid","Mike Johnson",
     date(2026,3,28),date(2026,4,30),date(2026,4,30),720000,"Bid Submitted",
     None,None,None,None,date(2026,5,25),0.60,
     date(2026,5,8),"Mike Johnson",date(2026,5,3),
     "Repeat client. Previous phase awarded to ASM. Strong position."),

    ("ASM-2026-006","Marriott Courtyard River District","DPR Construction","Richmond, VA",
     "Hospitality","Full Mechanical","GMP","GMP","Sarah Chen",
     date(2026,1,5),date(2026,2,10),date(2026,2,9),1450000,"Awarded",
     1450000,None,None,None,1.0,
     None,None,date(2026,3,20),
     "Awarded. In preconstruction. Construction start July 2026."),

    ("ASM-2026-007","Tech Campus Phase 1","Skanska USA","Austin, TX",
     "Commercial","Design-Build","Design-Build","Design-Build","Lisa Park",
     date(2026,4,1),date(2026,5,20),None,2200000,"In Progress",
     None,None,None,None,date(2026,7,1),0.30,
     date(2026,5,14),"Lisa Park",date(2026,4,15),
     "Design-build RFP. Partnering with MEP engineer of record."),

    ("ASM-2026-008","Riverside Student Housing","Clark Construction","Columbus, OH",
     "Student Housing","Plumbing Only","Hard Bid","Hard Bid","Tom Rodriguez",
     date(2026,2,14),date(2026,3,25),date(2026,3,24),410000,"Lost",
     None,"ABC Plumbing",385000,
     "Relationship",None,0,
     None,None,date(2026,4,2),
     "GC had prior relationship with ABC. Keep pursuing this GC."),

    ("ASM-2026-009","Summit Office Tower MEP Package","Hensel Phelps","Washington, DC",
     "Commercial","Full Mechanical","Permit Set","Hard Bid","Dave Williams",
     date(2026,4,5),date(2026,5,8),date(2026,5,7),890000,"Bid Submitted",
     None,None,None,None,date(2026,6,10),0.65,
     date(2026,5,13),"Dave Williams",date(2026,5,5),
     "Strong drawings. Labor availability risk in DC. Confirm pipe pricing."),

    ("ASM-2026-010","Lakewood Apartment Homes","New South Construction","Birmingham, AL",
     "Multifamily","Plumbing Only","Hard Bid","Hard Bid","Mike Johnson",
     date(2026,1,20),date(2026,2,28),date(2026,2,27),290000,"Awarded",
     295000,None,None,None,1.0,
     None,None,date(2026,3,10),
     "Awarded at buyout. Scope addition added $5K to original bid."),

    ("ASM-2026-011","Airport Hotel Renovation","Balfour Beatty","Houston, TX",
     "Hospitality","HVAC Only","Budget","Negotiated","Rachel Nguyen",
     date(2026,4,10),date(2026,5,2),date(2026,5,1),150000,"Budget Submitted",
     None,None,None,None,date(2026,6,15),0.40,
     date(2026,5,16),"Rachel Nguyen",date(2026,5,1),
     "Budget pricing only. Full hard bid expected in July 2026."),

    ("ASM-2026-012","Metro Medical Plaza","McCarthy Building","St. Louis, MO",
     "Healthcare","Full Mechanical","DD","GMP","James Carter",
     date(2026,3,10),date(2026,4,18),date(2026,4,17),1800000,"Bid Submitted",
     None,None,None,None,date(2026,5,30),0.45,
     date(2026,5,7),"James Carter",date(2026,4,30),
     "Complex healthcare. Infection control zones. Tight 18-month schedule."),

    ("ASM-2026-013","Eastside Commons Apartments","Choate Construction","Greenville, SC",
     "Multifamily","Full Mechanical","Hard Bid","Hard Bid","Sarah Chen",
     date(2026,4,15),date(2026,5,15),None,630000,"In Progress",
     None,None,None,None,date(2026,6,20),0.55,
     date(2026,5,19),"Sarah Chen",date(2026,4,20),
     "Drawings received. Quantity takeoff in progress."),

    ("ASM-2026-014","Central High School Mechanical Upgrade","Turner Construction","Louisville, KY",
     "Education","Plumbing Only","Hard Bid","Hard Bid","Tom Rodriguez",
     date(2026,2,1),date(2026,3,10),date(2026,3,9),380000,"Lost",
     None,"Metro Plumbing Co.",345000,
     "Price",None,0,
     None,None,date(2026,3,20),
     "Public school bid. Lost by $35K. Review material and labor pricing."),

    ("ASM-2026-015","Innovation Hub Full MEP","Webcor Builders","San Francisco, CA",
     "Commercial","Design-Build","Design-Build","Design-Build","Lisa Park",
     date(2026,4,20),date(2026,6,1),None,3100000,"In Progress",
     None,None,None,None,date(2026,8,1),0.25,
     date(2026,5,30),"Lisa Park",date(2026,4,25),
     "Large D-B opportunity. Confirm MEP engineer partner before submitting."),
]

PENDING_STATUSES = '"Bid Submitted","Budget Submitted","ROM Submitted","In Progress"'

def build_bid_log(ws):
    ws.sheet_properties.tabColor = BLUE
    ws.freeze_panes = "B2"
    ws.row_dimensions[1].height = 40

    for i, (name, width, _) in enumerate(BID_COLS):
        col = i + 1
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = width
        hcell(ws.cell(row=1, column=col), name, size=10, wrap=True)

    num_rows = len(SAMPLE_BIDS)
    for ri, bid in enumerate(SAMPLE_BIDS, start=2):
        ws.row_dimensions[ri].height = 18
        bg = LGRAY if ri % 2 == 0 else WHITE
        (bid_id, proj, gc, loc, mkt, scope, stage, deliv, est,
         recv, due, subm, bid_amt, status, award_amt, win_comp, comp_amt,
         lost_rsn, exp_award, prob, next_fu, fu_owner, last_ct, notes) = bid

        row_vals = [
            bid_id, proj, gc, loc, mkt, scope, stage, deliv, est,
            recv, due, subm, bid_amt, status, award_amt, win_comp, comp_amt,
            lost_rsn, exp_award, prob, next_fu, fu_owner, last_ct, notes,
            None, None,  # Go/No-Go Score, Rec (filled from GNG tab via VLOOKUP)
        ]
        for ci, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=ri, column=ci)
            if val is not None and val != "":
                cell.value = val
            cell.font  = F(10)
            cell.fill  = BG(bg)
            cell.alignment = AL("left", "center", ci == 24)  # wrap Notes
            cell.border = BD()
            _, _, fmt = BID_COLS[ci - 1]
            if fmt and fmt != "@":
                cell.number_format = fmt

    # ── Formula columns (27=AA, 28=AB, 29=AC) ────────────────────────────────
    for ri in range(2, num_rows + 2):
        n = str(ri)
        sref  = f"N{n}"   # Status
        lref  = f"L{n}"   # Submitted Date
        kref  = f"K{n}"   # Bid Due Date
        uref  = f"U{n}"   # Next Follow-Up
        vref  = f"V{n}"   # Follow-Up Owner
        bref  = f"B{n}"   # Project Name
        cref  = f"C{n}"   # GC
        iref  = f"I{n}"   # Estimator
        mref  = f"M{n}"   # ASM Bid Amount
        oref  = f"O{n}"   # Awarded Amount
        rref  = f"R{n}"   # Lost Reason
        bg = LGRAY if ri % 2 == 0 else WHITE

        # Col 25 (Y): Go/No-Go Score — VLOOKUP from GNG tab
        c25 = ws.cell(row=ri, column=25)
        c25.value = f'=IFERROR(VLOOKUP(A{n},\'Go-No-Go\'!$A:$M,12,0),"")'
        c25.font  = F(10, color=DGRAY); c25.fill = BG(bg)
        c25.alignment = AL("center"); c25.border = BD()
        c25.number_format = "0"

        # Col 26 (Z): Go/No-Go Rec — VLOOKUP from GNG tab
        c26 = ws.cell(row=ri, column=26)
        c26.value = f'=IFERROR(VLOOKUP(A{n},\'Go-No-Go\'!$A:$N,14,0),"")'
        c26.font  = F(10, color=DGRAY); c26.fill = BG(bg)
        c26.alignment = AL("center"); c26.border = BD()

        # Col 27 (AA): Days Pending
        c27 = ws.cell(row=ri, column=27)
        c27.value = (
            f'=IF(OR({sref}="Bid Submitted",{sref}="Budget Submitted",'
            f'{sref}="ROM Submitted",{sref}="In Progress"),'
            f'IF({lref}<>"",TODAY()-{lref},'
            f'IF({kref}<>"",TODAY()-{kref},"")),"")'
        )
        c27.font = F(10); c27.fill = BG(bg)
        c27.alignment = AL("center"); c27.border = BD()
        c27.number_format = "0"

        # Col 28 (AB): Follow-Up Status
        c28 = ws.cell(row=ri, column=28)
        pend = f'OR({sref}="Bid Submitted",{sref}="Budget Submitted",{sref}="ROM Submitted",{sref}="In Progress")'
        c28.value = (
            f'=IF(AND({pend},{uref}<>"",{uref}<TODAY()),"OVERDUE",'
            f'IF(AND({pend},{uref}<>"",{uref}<=TODAY()+3),"DUE SOON",'
            f'IF(AND({pend},{uref}=""),"MISSING","OK")))'
        )
        c28.font = F(10); c28.fill = BG(bg)
        c28.alignment = AL("center"); c28.border = BD()

        # Col 29 (AC): Data Quality Flag
        c29 = ws.cell(row=ri, column=29)
        c29.value = (
            f'=IFERROR(TRIM('
            f'IF({bref}="","[Missing Project Name] ","")&'
            f'IF({cref}="","[Missing GC] ","")&'
            f'IF({iref}="","[Missing Estimator] ","")&'
            f'IF({kref}="","[Missing Bid Due Date] ","")&'
            f'IF({mref}="","[Missing Bid Amount] ","")&'
            f'IF(AND({sref}="Awarded",{oref}=""),"[Awarded - No Award Amount] ","")&'
            f'IF(AND({sref}="Lost",{rref}=""),"[Lost - No Lost Reason] ","")&'
            f'IF(AND(OR({sref}="Bid Submitted",{sref}="Budget Submitted"),'
            f'{uref}=""),"[Pending - No Follow-Up Date] ","")&'
            f'IF(AND(OR({sref}="Bid Submitted",{sref}="Budget Submitted"),'
            f'"{vref}"=""),"[Pending - No Follow-Up Owner] ","")&'
            f'IF(AND({kref}<>"",{kref}<TODAY(),'
            f'OR({sref}="In Progress",{sref}="")),"[Bid Due Date Passed] ","")'
            f'),"OK")'
        )
        c29.font = F(9); c29.fill = BG(bg)
        c29.alignment = AL("left", "center", True); c29.border = BD()

    # ── Data Validation (dropdowns) ───────────────────────────────────────────
    last_row = 1001

    def add_dv(col_name, s_col, s_start, s_count):
        s_end = s_start + s_count - 1
        dv = DataValidation(
            type="list",
            formula1=f"Settings!${s_col}${s_start}:${s_col}${s_end}",
            allow_blank=True, showDropDown=False,
            showErrorMessage=True, errorTitle="Invalid Entry",
            error="Select a value from the dropdown list."
        )
        ws.add_data_validation(dv)
        c = cl(col_name)
        dv.sqref = f"{c}2:{c}{last_row}"

    # Settings rows: header=row4, values start row5
    add_dv("General Contractor", "Q", 5, 20)
    add_dv("Market Segment",     "E", 5, 10)
    add_dv("Scope Type",         "G", 5, 6)
    add_dv("Bid Stage",          "I", 5, 7)
    add_dv("Delivery Method",    "K", 5, 6)
    add_dv("Estimator",          "A", 5, 7)
    add_dv("Status",             "C", 5, 8)
    add_dv("Lost Reason",        "M", 5, 9)
    add_dv("Follow-Up Owner",    "S", 5, 9)

    # ── Conditional Formatting ────────────────────────────────────────────────
    data_rng  = f"A2:AC{last_row}"
    n_col     = cl("Status")           # N
    ab_col    = cl("Follow-Up Status") # AB
    aa_col    = cl("Days Pending")     # AA
    ac_col    = cl("Data Quality Flag")# AC

    def diff_style(font_color=None, fill_color=None, bold=False):
        kwargs = {}
        if font_color: kwargs["font"] = Font(name="Arial", size=10, bold=bold, color=font_color)
        if fill_color: kwargs["fill"] = PatternFill("solid", fgColor=fill_color)
        return DifferentialStyle(**kwargs)

    # Awarded rows → light green
    ws.conditional_formatting.add(data_rng, FormulaRule(
        formula=[f'${n_col}2="Awarded"'],
        dxf=diff_style(DKGRN, LTGRN), stopIfTrue=False))

    # Lost rows → light red
    ws.conditional_formatting.add(data_rng, FormulaRule(
        formula=[f'${n_col}2="Lost"'],
        dxf=diff_style(DKRED, LTRED), stopIfTrue=False))

    # Follow-Up OVERDUE → orange
    ws.conditional_formatting.add(f"{ab_col}2:{ab_col}{last_row}", FormulaRule(
        formula=[f'{ab_col}2="OVERDUE"'],
        dxf=diff_style("7F3300", "FFCC66", True), stopIfTrue=False))

    # Follow-Up DUE SOON → yellow
    ws.conditional_formatting.add(f"{ab_col}2:{ab_col}{last_row}", FormulaRule(
        formula=[f'{ab_col}2="DUE SOON"'],
        dxf=diff_style(DKYEL, LTYEL, True), stopIfTrue=False))

    # Follow-Up MISSING → light purple
    ws.conditional_formatting.add(f"{ab_col}2:{ab_col}{last_row}", FormulaRule(
        formula=[f'{ab_col}2="MISSING"'],
        dxf=diff_style(PURPLE, LTPURP, True), stopIfTrue=False))

    # Days Pending > 60 → red
    ws.conditional_formatting.add(f"{aa_col}2:{aa_col}{last_row}", FormulaRule(
        formula=[f'AND({aa_col}2>60,{aa_col}2<>"")'],
        dxf=diff_style(DKRED, LTRED, True), stopIfTrue=False))

    # Days Pending 31-60 → yellow
    ws.conditional_formatting.add(f"{aa_col}2:{aa_col}{last_row}", FormulaRule(
        formula=[f'AND({aa_col}2>30,{aa_col}2<=60,{aa_col}2<>"")'],
        dxf=diff_style(DKYEL, LTYEL, True), stopIfTrue=False))

    # Data Quality not OK → light red
    ws.conditional_formatting.add(f"{ac_col}2:{ac_col}{last_row}", FormulaRule(
        formula=[f'{ac_col}2<>"OK"'],
        dxf=diff_style(DKRED, "FFE0E0"), stopIfTrue=False))

    # Go/No-Go Rec coloring
    gng_col = cl("Go/No-Go Rec")  # Z
    ws.conditional_formatting.add(f"{gng_col}2:{gng_col}{last_row}", FormulaRule(
        formula=[f'{gng_col}2="Go"'],
        dxf=diff_style(DKGRN, LTGRN, True), stopIfTrue=False))
    ws.conditional_formatting.add(f"{gng_col}2:{gng_col}{last_row}", FormulaRule(
        formula=[f'{gng_col}2="No-Go"'],
        dxf=diff_style(DKRED, LTRED, True), stopIfTrue=False))
    ws.conditional_formatting.add(f"{gng_col}2:{gng_col}{last_row}", FormulaRule(
        formula=[f'{gng_col}2="Review"'],
        dxf=diff_style(DKYEL, LTYEL, True), stopIfTrue=False))

    # ── Excel Table ───────────────────────────────────────────────────────────
    tbl = Table(displayName="BidLog", ref=f"A1:AC{num_rows + 1}")
    tbl.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    ws.add_table(tbl)

# ══════════════════════════════════════════════════════════════════════════════
# GO / NO-GO TAB
# ══════════════════════════════════════════════════════════════════════════════
GNG_COLS = [
    ("Bid ID",                   12, "@"),
    ("Project Name",             30, "@"),
    ("GC",                       20, "@"),
    ("GC Relationship\n(1-5)",   13, "0"),
    ("Project Fit\n(1-5)",       13, "0"),
    ("Drawing Quality\n(1-5)",   13, "0"),
    ("Schedule Risk\n(1-5)",     13, "0"),
    ("Labor Availability\n(1-5)",14, "0"),
    ("Margin Potential\n(1-5)",  14, "0"),
    ("Strategic Value\n(1-5)",   14, "0"),
    ("Competition\n(1-5)",       13, "0"),
    ("Total Score\n(/40)",       14, "0"),
    ("Score %",                  11, "0%"),
    ("Recommendation",           16, "@"),
]

GNG_TIPS = [
    "5 = Strong relationship / partner GC",
    "5 = Ideal project type & size for ASM",
    "5 = Complete & clear drawings/specs",
    "5 = Low risk (flexible schedule)",
    "5 = Crew fully available",
    "5 = Strong margin opportunity",
    "5 = High strategic importance to ASM",
    "5 = Low competition / sole-source",
]

# Pre-scored sample bids (Bid ID, scores D-K)
GNG_SCORES = [
    ("ASM-2026-001", 4,5,4,4,4,4,4,3),
    ("ASM-2026-003", 3,5,3,3,4,4,5,3),
    ("ASM-2026-005", 5,5,4,4,5,4,5,4),
    ("ASM-2026-007", 2,4,3,3,3,4,5,2),
    ("ASM-2026-009", 3,4,5,4,3,4,3,3),
    ("ASM-2026-011", 3,3,2,4,4,3,3,4),
    ("ASM-2026-012", 4,4,3,2,3,5,5,3),
    ("ASM-2026-013", 3,4,3,4,4,4,4,4),
    ("ASM-2026-015", 2,5,2,2,3,5,5,2),
]

def build_gonogo(ws):
    ws.sheet_properties.tabColor = "C55A11"
    ws.freeze_panes = "D2"
    ws.row_dimensions[1].height = 50

    for i, (name, width, fmt) in enumerate(GNG_COLS):
        col = i + 1
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = width
        c = ws.cell(row=1, column=col)
        hcell(c, name, bg=NAVY, size=9, wrap=True)
        c.alignment = AL("center", "center", True)

    # Scoring guide in row 2 (light background)
    ws.row_dimensions[2].height = 14
    ws.merge_cells("A2:C2")
    c = ws["A2"]
    c.value = "SCORING GUIDE →"
    c.font  = F(8, True, DGRAY)
    c.fill  = BG("FFF2CC"); c.alignment = AL("center", "center")

    for i, tip in enumerate(GNG_TIPS):
        c = ws.cell(row=2, column=4 + i)
        c.value = tip
        c.font  = F(7, False, DGRAY, True)
        c.fill  = BG("FFF2CC")
        c.alignment = AL("center", "center", True)

    for ri, (bid_id, *scores) in enumerate(GNG_SCORES, start=3):
        ws.row_dimensions[ri].height = 18
        bg = LGRAY if ri % 2 == 0 else WHITE
        n = str(ri)

        # Bid ID
        c = ws.cell(row=ri, column=1)
        c.value = bid_id; c.font = F(10); c.fill = BG(bg)
        c.alignment = AL("center"); c.border = BD()

        # Project Name (VLOOKUP from Bid Log)
        c = ws.cell(row=ri, column=2)
        c.value = f"=IFERROR(VLOOKUP(A{n},'Bid Log'!$A:$B,2,0),\"\")"
        c.font = F(10); c.fill = BG(bg); c.alignment = AL("left"); c.border = BD()

        # GC (VLOOKUP)
        c = ws.cell(row=ri, column=3)
        c.value = f"=IFERROR(VLOOKUP(A{n},'Bid Log'!$A:$C,3,0),\"\")"
        c.font = F(10); c.fill = BG(bg); c.alignment = AL("left"); c.border = BD()

        # Score columns D-K (cols 4-11)
        for ci, score in enumerate(scores, start=4):
            c = ws.cell(row=ri, column=ci)
            c.value = score
            c.font = F(10, True); c.fill = BG(bg)
            c.alignment = AL("center"); c.border = BD()
            c.number_format = "0"

        # Total Score (col 12)
        c = ws.cell(row=ri, column=12)
        c.value = f"=SUM(D{n}:K{n})"
        c.font = F(10, True); c.fill = BG(bg)
        c.alignment = AL("center"); c.border = BD()
        c.number_format = "0"

        # Score % (col 13)
        c = ws.cell(row=ri, column=13)
        c.value = f"=IF(L{n}>0,L{n}/40,\"\")"
        c.font = F(10, True); c.fill = BG(bg)
        c.alignment = AL("center"); c.border = BD()
        c.number_format = "0%"

        # Recommendation (col 14)
        c = ws.cell(row=ri, column=14)
        c.value = (
            f'=IF(M{n}="","",IF(M{n}>=0.70,"Go",'
            f'IF(M{n}>=0.50,"Review","No-Go")))'
        )
        c.font = F(10, True); c.fill = BG(bg)
        c.alignment = AL("center"); c.border = BD()

    last_gng = len(GNG_SCORES) + 3

    # Conditional formatting on Recommendation column (N=14)
    rec_range = f"N3:N{last_gng}"
    ws.conditional_formatting.add(rec_range, FormulaRule(
        formula=['N3="Go"'], dxf=DifferentialStyle(
            font=Font(name="Arial", size=10, bold=True, color=DKGRN),
            fill=PatternFill("solid", fgColor=LTGRN)), stopIfTrue=False))
    ws.conditional_formatting.add(rec_range, FormulaRule(
        formula=['N3="No-Go"'], dxf=DifferentialStyle(
            font=Font(name="Arial", size=10, bold=True, color=DKRED),
            fill=PatternFill("solid", fgColor=LTRED)), stopIfTrue=False))
    ws.conditional_formatting.add(rec_range, FormulaRule(
        formula=['N3="Review"'], dxf=DifferentialStyle(
            font=Font(name="Arial", size=10, bold=True, color=DKYEL),
            fill=PatternFill("solid", fgColor=LTYEL)), stopIfTrue=False))

    # Score % color scale (col M)
    score_range = f"M3:M{last_gng}"
    ws.conditional_formatting.add(score_range, FormulaRule(
        formula=['AND(M3>=0.70,M3<>"")'], dxf=DifferentialStyle(
            fill=PatternFill("solid", fgColor=LTGRN)), stopIfTrue=False))
    ws.conditional_formatting.add(score_range, FormulaRule(
        formula=['AND(M3<0.50,M3<>"")'], dxf=DifferentialStyle(
            fill=PatternFill("solid", fgColor=LTRED)), stopIfTrue=False))

    # Score cells 1-2 red, 4-5 green
    score_rng = f"D3:K{last_gng}"
    ws.conditional_formatting.add(score_rng, FormulaRule(
        formula=['D3<=2'], dxf=DifferentialStyle(
            fill=PatternFill("solid", fgColor="FFD0D0")), stopIfTrue=False))
    ws.conditional_formatting.add(score_rng, FormulaRule(
        formula=['D3>=4'], dxf=DifferentialStyle(
            fill=PatternFill("solid", fgColor="D0F0D0")), stopIfTrue=False))

    # Add scoring explanation note below table
    note_row = last_gng + 2
    ws.row_dimensions[note_row].height = 20
    ws.merge_cells(f"A{note_row}:N{note_row}")
    c = ws[f"A{note_row}"]
    c.value = "SCORING KEY:  Each category scored 1–5.  Max score = 40.  |  GO = 70%+ (score ≥ 28)   |   REVIEW = 50–69% (score 20–27)   |   NO-GO = below 50% (score < 20)"
    c.font  = F(9, True, DGRAY)
    c.fill  = BG("FFF2CC")
    c.alignment = AL("left", "center")

# ══════════════════════════════════════════════════════════════════════════════
# DASHBOARD TAB
# ══════════════════════════════════════════════════════════════════════════════
def build_dashboard(ws):
    ws.sheet_properties.tabColor = "375623"
    ws.sheet_view.showGridLines = False

    # Set column widths
    col_widths = [2, 20, 20, 2, 20, 20, 2, 20, 20, 2, 20, 20, 2, 20, 20, 2]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Title ──────────────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 36
    ws.merge_cells("B1:P1")
    c = ws["B1"]
    c.value = "ASM PRECONSTRUCTION PIPELINE TRACKER V9  —  DASHBOARD"
    c.font  = Font(name="Arial", size=16, bold=True, color=WHITE)
    c.fill  = BG(NAVY)
    c.alignment = AL("center", "center")

    ws.row_dimensions[2].height = 18
    ws.merge_cells("B2:P2")
    c = ws["B2"]
    c.value = f'="All States Mechanical  |  Data as of: "&TEXT(TODAY(),"mmmm d, yyyy")'
    c.font  = F(10, False, DGRAY)
    c.alignment = AL("center", "center")

    # ── KPI Row 1 — Volume ──────────────────────────────────────────────────
    def kpi_box(label_cell, val_cell, merge_label, merge_val, label, formula,
                bg, num_fmt="$#,##0", val_size=20):
        ws.row_dimensions[int(merge_label[1:])].height = 18
        ws.row_dimensions[int(merge_val[1:])].height = 32
        ws.merge_cells(merge_label)
        lc = ws[merge_label.split(":")[0]]
        lc.value = label
        lc.font  = Font(name="Arial", size=9, bold=True, color=WHITE)
        lc.fill  = BG(bg)
        lc.alignment = AL("center", "center", True)
        ws.merge_cells(merge_val)
        vc = ws[merge_val.split(":")[0]]
        vc.value = formula
        vc.font  = Font(name="Arial", size=val_size, bold=True, color=WHITE)
        vc.fill  = BG(bg)
        vc.alignment = AL("center", "center")
        vc.number_format = num_fmt

    # Row 3 labels, Row 4 values
    kpi_box("B3","B4", "B3:C3","B4:C4",
        "TOTAL BID VOLUME YTD",
        '=IFERROR(SUMIF(\'Bid Log\'!N:N,"<>No Bid",\'Bid Log\'!M:M),0)',
        NAVY)
    kpi_box("E3","E4", "E3:F3","E4:F4",
        "AWARDED VOLUME YTD",
        '=IFERROR(SUMIF(\'Bid Log\'!N:N,"Awarded",\'Bid Log\'!O:O),0)',
        DKGRN)
    kpi_box("H3","H4", "H3:I3","H4:I4",
        "LOST VOLUME YTD",
        '=IFERROR(SUMIF(\'Bid Log\'!N:N,"Lost",\'Bid Log\'!M:M),0)',
        DKRED)
    kpi_box("K3","K4", "K3:L3","K4:L4",
        "PENDING VOLUME",
        '=IFERROR(SUMIFS(\'Bid Log\'!M:M,\'Bid Log\'!N:N,"Bid Submitted")+SUMIFS(\'Bid Log\'!M:M,\'Bid Log\'!N:N,"Budget Submitted")+SUMIFS(\'Bid Log\'!M:M,\'Bid Log\'!N:N,"ROM Submitted"),0)',
        BLUE)
    kpi_box("N3","N4", "N3:O3","N4:O4",
        "WIN RATE (BY $)",
        '=IFERROR(SUMIF(\'Bid Log\'!N:N,"Awarded",\'Bid Log\'!O:O)/(SUMIF(\'Bid Log\'!N:N,"Awarded",\'Bid Log\'!O:O)+SUMIF(\'Bid Log\'!N:N,"Lost",\'Bid Log\'!M:M)),0)',
        TEAL, num_fmt="0.0%", val_size=22)

    ws.row_dimensions[5].height = 8  # spacer

    # ── KPI Row 2 — Activity ────────────────────────────────────────────────
    kpi_box("B6","B7", "B6:C6","B7:C7",
        "PENDING HIGH-PROB (≥70%)",
        '=IFERROR(SUMPRODUCT((OR(\'Bid Log\'!N2:N1001="Bid Submitted",\'Bid Log\'!N2:N1001="Budget Submitted"))*(\'Bid Log\'!T2:T1001>=0.7)*\'Bid Log\'!M2:M1001),0)',
        "1F6B75")
    kpi_box("E6","E7", "E6:F6","E7:F7",
        "BIDS DUE NEXT 7 DAYS",
        '=IFERROR(COUNTIFS(\'Bid Log\'!K:K,">="&TODAY(),\'Bid Log\'!K:K,"<="&TODAY()+7,\'Bid Log\'!N:N,"In Progress"),0)',
        ORANGE, num_fmt="0", val_size=24)
    kpi_box("H6","H7", "H6:I6","H7:I7",
        "FOLLOW-UPS PAST DUE",
        '=IFERROR(COUNTIF(\'Bid Log\'!AB:AB,"OVERDUE"),0)',
        DKRED, num_fmt="0", val_size=24)
    kpi_box("K6","K7", "K6:L6","K7:L7",
        "EXPECTED AWARDS THIS MONTH",
        '=IFERROR(COUNTIFS(\'Bid Log\'!S:S,">="&DATE(YEAR(TODAY()),MONTH(TODAY()),1),\'Bid Log\'!S:S,"<="&EOMONTH(TODAY(),0),\'Bid Log\'!N:N,"Bid Submitted"),0)',
        PURPLE, num_fmt="0", val_size=24)
    kpi_box("N6","N7", "N6:O6","N7:O7",
        "WIN RATE (BY COUNT)",
        '=IFERROR(COUNTIF(\'Bid Log\'!N:N,"Awarded")/(COUNTIF(\'Bid Log\'!N:N,"Awarded")+COUNTIF(\'Bid Log\'!N:N,"Lost")),0)',
        "1F6B75", num_fmt="0.0%", val_size=22)

    ws.row_dimensions[8].height = 10  # spacer

    # ── Section: Bid Volume by Market Segment ──────────────────────────────
    r = 9
    section_bar(ws, r, "B", "O", "  BID VOLUME BY MARKET SEGMENT", bg=NAVY)

    headers = ["Market Segment", "# Bids", "Bid Volume", "# Awarded", "Awarded Volume", "Win Rate %"]
    h_cols  = ["B","D","F","H","J","L"]
    h_widths = [20, 8, 18, 10, 18, 10]
    for col_letter, hdr_text, w in zip(h_cols, headers, h_widths):
        ws.column_dimensions[col_letter].width = w
        hcell(ws.cell(row=r+1, column=column_index_from_string(col_letter)),
              hdr_text, bg=BLUE, size=9)
        ws.row_dimensions[r+1].height = 18

    segs, _ = zip(*[(s[0], None) for s in SETTINGS_LISTS if s[0] == "MARKET SEGMENTS"])
    market_segs = [s for s in SETTINGS_LISTS if s[0] == "MARKET SEGMENTS"][0][2]

    for i, seg in enumerate(market_segs):
        rr = r + 2 + i
        ws.row_dimensions[rr].height = 16
        bg = LGRAY if i % 2 == 0 else WHITE
        seg_q = f'"{seg}"'
        vcell(ws.cell(rr, column_index_from_string("B")), seg, bg=bg, size=9)
        ws.cell(rr, column_index_from_string("D")).value = f'=COUNTIF(\'Bid Log\'!E:E,{seg_q})'
        ws.cell(rr, column_index_from_string("F")).value = f'=IFERROR(SUMIFS(\'Bid Log\'!M:M,\'Bid Log\'!E:E,{seg_q},\'Bid Log\'!N:N,"<>No Bid"),0)'
        ws.cell(rr, column_index_from_string("H")).value = f'=COUNTIFS(\'Bid Log\'!E:E,{seg_q},\'Bid Log\'!N:N,"Awarded")'
        ws.cell(rr, column_index_from_string("J")).value = f'=IFERROR(SUMIFS(\'Bid Log\'!O:O,\'Bid Log\'!E:E,{seg_q},\'Bid Log\'!N:N,"Awarded"),0)'
        ws.cell(rr, column_index_from_string("L")).value = f'=IFERROR(H{rr}/D{rr},"")'
        for col_l in ["B","D","F","H","J","L"]:
            cell = ws.cell(rr, column_index_from_string(col_l))
            cell.font  = F(9)
            cell.fill  = BG(bg)
            cell.alignment = AL("left" if col_l=="B" else "center")
            cell.border = BD()
        ws.cell(rr, column_index_from_string("F")).number_format = "$#,##0"
        ws.cell(rr, column_index_from_string("J")).number_format = "$#,##0"
        ws.cell(rr, column_index_from_string("L")).number_format = "0%"

    next_row = r + 2 + len(market_segs) + 1

    # ── Section: Estimator Workload ─────────────────────────────────────────
    ws.row_dimensions[next_row].height = 10
    r2 = next_row + 1
    section_bar(ws, r2, "B", "O", "  ESTIMATOR WORKLOAD & PERFORMANCE", bg=NAVY)
    est_headers = ["Estimator", "Active Bids", "Active Volume", "Awarded", "Awarded $", "Lost", "Win Rate"]
    est_cols    = ["B","D","F","H","J","L","N"]
    for col_letter, hdr_text in zip(est_cols, est_headers):
        hcell(ws.cell(row=r2+1, column=column_index_from_string(col_letter)),
              hdr_text, bg=BLUE, size=9)
        ws.row_dimensions[r2+1].height = 18

    estimators = [s for s in SETTINGS_LISTS if s[0] == "ESTIMATORS"][0][2]
    for i, est in enumerate(estimators):
        rr = r2 + 2 + i
        ws.row_dimensions[rr].height = 16
        bg = LGRAY if i % 2 == 0 else WHITE
        eq = f'"{est}"'
        active = f'COUNTIFS(\'Bid Log\'!I:I,{eq},\'Bid Log\'!N:N,"Bid Submitted")+COUNTIFS(\'Bid Log\'!I:I,{eq},\'Bid Log\'!N:N,"Budget Submitted")+COUNTIFS(\'Bid Log\'!I:I,{eq},\'Bid Log\'!N:N,"In Progress")'
        vcell(ws.cell(rr, column_index_from_string("B")), est, bg=bg, size=9)
        ws.cell(rr, column_index_from_string("D")).value = f'={active}'
        ws.cell(rr, column_index_from_string("F")).value = (
            f'=IFERROR(SUMIFS(\'Bid Log\'!M:M,\'Bid Log\'!I:I,{eq},'
            f'\'Bid Log\'!N:N,"Bid Submitted")+SUMIFS(\'Bid Log\'!M:M,\'Bid Log\'!I:I,{eq},'
            f'\'Bid Log\'!N:N,"Budget Submitted"),0)')
        ws.cell(rr, column_index_from_string("H")).value = f'=COUNTIFS(\'Bid Log\'!I:I,{eq},\'Bid Log\'!N:N,"Awarded")'
        ws.cell(rr, column_index_from_string("J")).value = f'=IFERROR(SUMIFS(\'Bid Log\'!O:O,\'Bid Log\'!I:I,{eq},\'Bid Log\'!N:N,"Awarded"),0)'
        ws.cell(rr, column_index_from_string("L")).value = f'=COUNTIFS(\'Bid Log\'!I:I,{eq},\'Bid Log\'!N:N,"Lost")'
        ws.cell(rr, column_index_from_string("N")).value = f'=IFERROR(H{rr}/(H{rr}+L{rr}),"")'
        for col_l in ["B","D","F","H","J","L","N"]:
            cell = ws.cell(rr, column_index_from_string(col_l))
            cell.font = F(9); cell.fill = BG(bg)
            cell.alignment = AL("left" if col_l=="B" else "center")
            cell.border = BD()
        ws.cell(rr, column_index_from_string("F")).number_format = "$#,##0"
        ws.cell(rr, column_index_from_string("J")).number_format = "$#,##0"
        ws.cell(rr, column_index_from_string("N")).number_format = "0%"

    r3_start = r2 + 2 + len(estimators) + 2

    # ── Section: Lost Reasons ───────────────────────────────────────────────
    section_bar(ws, r3_start, "B", "O", "  LOSSES BY REASON", bg=NAVY)
    hcell(ws.cell(r3_start+1, column_index_from_string("B")), "Lost Reason", bg=BLUE, size=9)
    hcell(ws.cell(r3_start+1, column_index_from_string("D")), "# Lost", bg=BLUE, size=9)
    hcell(ws.cell(r3_start+1, column_index_from_string("F")), "Lost Volume", bg=BLUE, size=9)
    ws.row_dimensions[r3_start+1].height = 18
    lost_reasons = [s for s in SETTINGS_LISTS if s[0] == "LOST REASONS"][0][2]
    for i, rsn in enumerate(lost_reasons):
        rr = r3_start + 2 + i
        ws.row_dimensions[rr].height = 16
        bg = LGRAY if i % 2 == 0 else WHITE
        rq = f'"{rsn}"'
        vcell(ws.cell(rr, column_index_from_string("B")), rsn, bg=bg, size=9)
        ws.cell(rr, column_index_from_string("D")).value = f'=COUNTIFS(\'Bid Log\'!R:R,{rq},\'Bid Log\'!N:N,"Lost")'
        ws.cell(rr, column_index_from_string("F")).value = f'=IFERROR(SUMIFS(\'Bid Log\'!M:M,\'Bid Log\'!R:R,{rq},\'Bid Log\'!N:N,"Lost"),0)'
        for col_l in ["B","D","F"]:
            cell = ws.cell(rr, column_index_from_string(col_l))
            cell.font = F(9); cell.fill = BG(bg)
            cell.alignment = AL("left" if col_l=="B" else "center")
            cell.border = BD()
        ws.cell(rr, column_index_from_string("F")).number_format = "$#,##0"

# ══════════════════════════════════════════════════════════════════════════════
# FRIDAY MEETING TAB
# ══════════════════════════════════════════════════════════════════════════════
def build_friday(ws):
    ws.sheet_properties.tabColor = PURPLE
    ws.sheet_view.showGridLines = False

    for i, w in enumerate([2,22,22,16,16,14,12,12,2], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    title_bar(ws, 1, "B", "I",
        '="ASM FRIDAY ESTIMATING MEETING  |  "&TEXT(TODAY(),"mmmm d, yyyy")',
        bg=NAVY)

    ws.row_dimensions[2].height = 22
    ws.merge_cells("B2:I2")
    c = ws["B2"]
    c.value = "Review each section below before the meeting. Go to the Bid Log tab and use Data > Filter to pull specific bid lists."
    c.font = F(10, False, DGRAY, True)
    c.alignment = AL("left", "center")

    r = 3
    def fri_section(title, bg_col):
        nonlocal r
        ws.row_dimensions[r].height = 10; r += 1
        section_bar(ws, r, "B", "I", f"  {title}", bg=bg_col)
        r += 1

    def metric_row(label, formula, num_fmt="0", bold_val=True):
        nonlocal r
        ws.row_dimensions[r].height = 18
        c1 = ws.cell(row=r, column=2)
        c1.value = label; c1.font = F(10); c1.fill = BG(LGRAY)
        c1.alignment = AL("left", "center"); c1.border = BD()
        ws.merge_cells(f"B{r}:D{r}")
        c2 = ws.cell(row=r, column=5)
        c2.value = formula
        c2.font = F(10, bold_val)
        c2.fill = BG(WHITE); c2.alignment = AL("center"); c2.border = BD()
        c2.number_format = num_fmt
        r += 1

    def note_row(text):
        nonlocal r
        ws.row_dimensions[r].height = 16
        ws.merge_cells(f"B{r}:I{r}")
        c = ws[f"B{r}"]
        c.value = text; c.font = F(9, False, DGRAY, True)
        c.fill = BG("FFFACD"); c.alignment = AL("left", "center")
        r += 1

    # ── BIDDING PIPELINE ────────────────────────────────────────────────────
    fri_section("1.  BIDS DUE IN THE NEXT 14 DAYS", BLUE)
    metric_row("Bids due in next 7 days (In Progress)",
        "=COUNTIFS('Bid Log'!K:K,\">=\"&TODAY(),'Bid Log'!K:K,\"<=\"&TODAY()+7,'Bid Log'!N:N,\"In Progress\")")
    metric_row("Bids due in days 8-14 (In Progress)",
        "=COUNTIFS('Bid Log'!K:K,\">=\"&TODAY()+8,'Bid Log'!K:K,\"<=\"&TODAY()+14,'Bid Log'!N:N,\"In Progress\")")
    metric_row("Total volume due in next 14 days",
        "=IFERROR(SUMIFS('Bid Log'!M:M,'Bid Log'!K:K,\">=\"&TODAY(),'Bid Log'!K:K,\"<=\"&TODAY()+14,'Bid Log'!N:N,\"In Progress\"),0)",
        num_fmt="$#,##0")
    note_row("▶  Filter Bid Log: Status = In Progress | Bid Due Date between today and today+14")

    # ── FOLLOW-UP ───────────────────────────────────────────────────────────
    fri_section("2.  FOLLOW-UP STATUS", DKRED)
    metric_row("Follow-ups OVERDUE",
        "=COUNTIF('Bid Log'!AB:AB,\"OVERDUE\")")
    metric_row("Follow-ups DUE THIS WEEK",
        "=COUNTIF('Bid Log'!AB:AB,\"DUE SOON\")")
    metric_row("Pending bids with NO follow-up date",
        "=COUNTIF('Bid Log'!AB:AB,\"MISSING\")")
    metric_row("Pending bids older than 30 days",
        "=COUNTIFS('Bid Log'!AA:AA,\">30\",'Bid Log'!AA:AA,\"<>\")")
    metric_row("Pending bids older than 60 days",
        "=COUNTIFS('Bid Log'!AA:AA,\">60\",'Bid Log'!AA:AA,\"<>\")")
    note_row("▶  Filter Bid Log: Follow-Up Status = OVERDUE  |  Sort by Next Follow-Up Date ascending")

    # ── EXPECTED AWARDS ──────────────────────────────────────────────────────
    fri_section("3.  EXPECTED AWARDS THIS MONTH", DKGRN)
    metric_row("Bids expecting award this month",
        "=COUNTIFS('Bid Log'!S:S,\">=\"&DATE(YEAR(TODAY()),MONTH(TODAY()),1),'Bid Log'!S:S,\"<=\"&EOMONTH(TODAY(),0),'Bid Log'!N:N,\"Bid Submitted\")")
    metric_row("Expected award volume this month",
        "=IFERROR(SUMPRODUCT((MONTH('Bid Log'!S2:S1001)=MONTH(TODAY()))*(YEAR('Bid Log'!S2:S1001)=YEAR(TODAY()))*('Bid Log'!N2:N1001=\"Bid Submitted\")*'Bid Log'!M2:M1001),0)",
        num_fmt="$#,##0")
    metric_row("High-probability pending (≥70%) count",
        "=IFERROR(SUMPRODUCT((('Bid Log'!N2:N1001=\"Bid Submitted\")+('Bid Log'!N2:N1001=\"Budget Submitted\")>0)*('Bid Log'!T2:T1001>=0.7)),0)")
    metric_row("High-probability pending (≥70%) volume",
        "=IFERROR(SUMPRODUCT((('Bid Log'!N2:N1001=\"Bid Submitted\")+('Bid Log'!N2:N1001=\"Budget Submitted\")>0)*('Bid Log'!T2:T1001>=0.7)*'Bid Log'!M2:M1001),0)",
        num_fmt="$#,##0")
    note_row("▶  Filter Bid Log: Status = Bid Submitted | Expected Award Date = this month  |  Sort by Probability descending")

    # ── GO / NO-GO ───────────────────────────────────────────────────────────
    fri_section("4.  GO / NO-GO ITEMS TO REVIEW", ORANGE)
    metric_row("Bids scored No-Go in Go/No-Go tab",
        "=COUNTIF('Bid Log'!Z:Z,\"No-Go\")")
    metric_row("Bids scored Review (decision needed)",
        "=COUNTIF('Bid Log'!Z:Z,\"Review\")")
    metric_row("Active bids with no Go/No-Go score",
        "=COUNTIFS('Bid Log'!N:N,\"In Progress\",'Bid Log'!Y:Y,\"\")")
    note_row("▶  Go to Go-No-Go tab to review scores  |  Filter Bid Log: Go/No-Go Rec = No-Go or Review")

    # ── AWARDS & LOSSES ─────────────────────────────────────────────────────
    fri_section("5.  RECENT AWARDS & LOSSES (LAST 30 DAYS)", TEAL)
    metric_row("Jobs awarded in last 30 days (count)",
        "=COUNTIFS('Bid Log'!W:W,\">=\"&TODAY()-30,'Bid Log'!N:N,\"Awarded\")")
    metric_row("Jobs awarded in last 30 days (volume)",
        "=IFERROR(SUMPRODUCT(('Bid Log'!W2:W1001>=TODAY()-30)*('Bid Log'!N2:N1001=\"Awarded\")*'Bid Log'!O2:O1001),0)",
        num_fmt="$#,##0")
    metric_row("Jobs lost in last 30 days (count)",
        "=COUNTIFS('Bid Log'!W:W,\">=\"&TODAY()-30,'Bid Log'!N:N,\"Lost\")")
    metric_row("Jobs lost in last 30 days (volume)",
        "=IFERROR(SUMPRODUCT(('Bid Log'!W2:W1001>=TODAY()-30)*('Bid Log'!N2:N1001=\"Lost\")*'Bid Log'!M2:M1001),0)",
        num_fmt="$#,##0")
    note_row("▶  Filter Bid Log: Status = Awarded or Lost | Last Contact Date >= today-30")

    # ── ESTIMATOR WORKLOAD ───────────────────────────────────────────────────
    fri_section("6.  ESTIMATOR WORKLOAD SNAPSHOT", NAVY)
    ws.row_dimensions[r].height = 18
    for ci, hdr_text in enumerate(["Estimator","Active Bids","Active Volume","Bids Due Next 14 Days"], start=2):
        hcell(ws.cell(r, ci+1), hdr_text, bg=BLUE, size=9)
    r += 1

    estimators = [s for s in SETTINGS_LISTS if s[0] == "ESTIMATORS"][0][2]
    for i, est in enumerate(estimators):
        ws.row_dimensions[r].height = 16
        bg = LGRAY if i % 2 == 0 else WHITE
        eq = f'"{est}"'
        vcell(ws.cell(r, 3), est, bg=bg, size=9)
        ws.cell(r, 4).value = (
            f'=COUNTIFS(\'Bid Log\'!I:I,{eq},\'Bid Log\'!N:N,"Bid Submitted")'
            f'+COUNTIFS(\'Bid Log\'!I:I,{eq},\'Bid Log\'!N:N,"Budget Submitted")'
            f'+COUNTIFS(\'Bid Log\'!I:I,{eq},\'Bid Log\'!N:N,"In Progress")'
        )
        ws.cell(r, 5).value = (
            f'=IFERROR(SUMIFS(\'Bid Log\'!M:M,\'Bid Log\'!I:I,{eq},'
            f'\'Bid Log\'!N:N,"Bid Submitted")+SUMIFS(\'Bid Log\'!M:M,\'Bid Log\'!I:I,{eq},'
            f'\'Bid Log\'!N:N,"Budget Submitted")+SUMIFS(\'Bid Log\'!M:M,\'Bid Log\'!I:I,{eq},'
            f'\'Bid Log\'!N:N,"In Progress"),0)'
        )
        ws.cell(r, 6).value = (
            f'=COUNTIFS(\'Bid Log\'!I:I,{eq},\'Bid Log\'!K:K,">="&TODAY(),'
            f'\'Bid Log\'!K:K,"<="&TODAY()+14)'
        )
        for ci in [3,4,5,6]:
            cell = ws.cell(r, ci)
            cell.font = F(9); cell.fill = BG(bg)
            cell.alignment = AL("left" if ci==3 else "center"); cell.border = BD()
        ws.cell(r, 5).number_format = "$#,##0"
        r += 1

    ws.row_dimensions[r].height = 10; r += 1
    ws.merge_cells(f"B{r}:I{r}")
    c = ws[f"B{r}"]
    c.value = "Recommended alert: Any estimator with 3+ active bids AND 2+ due in next 14 days is at capacity. Discuss workload rebalancing at meeting."
    c.font = F(9, True, ORANGE); c.fill = BG(LTORANG)
    c.alignment = AL("left", "center", True)
    ws.row_dimensions[r].height = 24

# ══════════════════════════════════════════════════════════════════════════════
# FOLLOW-UP CONTROL TAB
# ══════════════════════════════════════════════════════════════════════════════
def build_followup(ws):
    ws.sheet_properties.tabColor = DKRED
    ws.sheet_view.showGridLines = False

    for i, w in enumerate([2,22,20,16,16,16,16,14,2], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    title_bar(ws, 1, "B", "H", "ASM — FOLLOW-UP CONTROL CENTER", bg=DKRED)

    ws.row_dimensions[2].height = 18
    ws.merge_cells("B2:H2")
    c = ws["B2"]
    c.value = "Use this tab to manage follow-up on pending bids. All data pulls from the Bid Log."
    c.font = F(10, False, DGRAY, True); c.alignment = AL("left","center")

    r = 3

    def fu_section(title):
        nonlocal r
        ws.row_dimensions[r].height = 8; r += 1
        section_bar(ws, r, "B", "H", f"  {title}", bg=DKRED); r += 1

    def fu_metric(label, formula, num_fmt="0"):
        nonlocal r
        ws.row_dimensions[r].height = 18
        ws.merge_cells(f"B{r}:D{r}")
        c1 = ws[f"B{r}"]
        c1.value = label; c1.font = F(10); c1.fill = BG(LGRAY)
        c1.alignment = AL("left","center"); c1.border = BD()
        c2 = ws.cell(r, 5)
        c2.value = formula; c2.font = F(11, True)
        c2.fill = BG(WHITE); c2.alignment = AL("center"); c2.border = BD()
        c2.number_format = num_fmt
        r += 1

    fu_section("FOLLOW-UP SUMMARY")
    fu_metric("Total pending bids (Bid/Budget/ROM Submitted)",
        "=COUNTIF('Bid Log'!N:N,\"Bid Submitted\")+COUNTIF('Bid Log'!N:N,\"Budget Submitted\")+COUNTIF('Bid Log'!N:N,\"ROM Submitted\")+COUNTIF('Bid Log'!N:N,\"In Progress\")")
    fu_metric("Follow-ups OVERDUE",
        "=COUNTIF('Bid Log'!AB:AB,\"OVERDUE\")")
    fu_metric("Follow-ups DUE WITHIN 3 DAYS",
        "=COUNTIF('Bid Log'!AB:AB,\"DUE SOON\")")
    fu_metric("Pending bids with NO follow-up date assigned",
        "=COUNTIF('Bid Log'!AB:AB,\"MISSING\")")
    fu_metric("Pending bids with NO follow-up owner assigned",
        "=COUNTIFS('Bid Log'!V:V,\"\",\'Bid Log'!N:N,\"Bid Submitted\")+COUNTIFS('Bid Log'!V:V,\"\",\'Bid Log'!N:N,\"Budget Submitted\")")

    fu_section("AGING ANALYSIS — DAYS SINCE SUBMITTED")
    fu_metric("Pending bids: 0–30 days old",
        "=COUNTIFS('Bid Log'!AA:AA,\">0\",'Bid Log'!AA:AA,\"<=30\")")
    fu_metric("Pending bids: 31–60 days old (ATTENTION)",
        "=COUNTIFS('Bid Log'!AA:AA,\">30\",'Bid Log'!AA:AA,\"<=60\")")
    fu_metric("Pending bids: over 60 days old (URGENT)",
        "=COUNTIFS('Bid Log'!AA:AA,\">60\",'Bid Log'!AA:AA,\"<>\")")
    fu_metric("Total pending volume: 0-30 days",
        "=IFERROR(SUMPRODUCT(('Bid Log'!AA2:AA1001>0)*('Bid Log'!AA2:AA1001<=30)*'Bid Log'!M2:M1001),0)",
        "$#,##0")
    fu_metric("Total pending volume: 31-60 days",
        "=IFERROR(SUMPRODUCT(('Bid Log'!AA2:AA1001>30)*('Bid Log'!AA2:AA1001<=60)*'Bid Log'!M2:M1001),0)",
        "$#,##0")
    fu_metric("Total pending volume: over 60 days",
        "=IFERROR(SUMPRODUCT(('Bid Log'!AA2:AA1001>60)*'Bid Log'!M2:M1001),0)",
        "$#,##0")

    fu_section("HIGH-VALUE PENDING BIDS (top threshold tracking)")
    fu_metric("Pending bids over $1,000,000",
        "=COUNTIFS('Bid Log'!M:M,\">1000000\",'Bid Log'!N:N,\"Bid Submitted\")+COUNTIFS('Bid Log'!M:M,\">1000000\",'Bid Log'!N:N,\"Budget Submitted\")")
    fu_metric("Total pending volume over $1M",
        "=IFERROR(SUMPRODUCT(('Bid Log'!M2:M1001>1000000)*(('Bid Log'!N2:N1001=\"Bid Submitted\")+('Bid Log'!N2:N1001=\"Budget Submitted\")>0)*'Bid Log'!M2:M1001),0)",
        "$#,##0")
    fu_metric("Bids over $500K with no follow-up date",
        "=IFERROR(SUMPRODUCT(('Bid Log'!M2:M1001>500000)*(('Bid Log'!N2:N1001=\"Bid Submitted\")+('Bid Log'!N2:N1001=\"Budget Submitted\")>0)*('Bid Log'!U2:U1001=\"\")),0)")

    fu_section("FOLLOW-UP BY OWNER")
    ws.row_dimensions[r].height = 18
    for ci, h in enumerate(["Follow-Up Owner","Overdue","Due Soon","Missing Date","Total Pending"], start=2):
        hcell(ws.cell(r, ci+1), h, bg=DKRED, size=9)
    r += 1

    fu_owners = [s for s in SETTINGS_LISTS if s[0] == "FOLLOW-UP OWNERS"][0][2]
    for i, owner in enumerate(fu_owners):
        ws.row_dimensions[r].height = 16
        bg = LGRAY if i % 2 == 0 else WHITE
        oq = f'"{owner}"'
        vcell(ws.cell(r, 3), owner, bg=bg, size=9)
        ws.cell(r, 4).value = f'=COUNTIFS(\'Bid Log\'!V:V,{oq},\'Bid Log\'!AB:AB,"OVERDUE")'
        ws.cell(r, 5).value = f'=COUNTIFS(\'Bid Log\'!V:V,{oq},\'Bid Log\'!AB:AB,"DUE SOON")'
        ws.cell(r, 6).value = f'=COUNTIFS(\'Bid Log\'!V:V,{oq},\'Bid Log\'!AB:AB,"MISSING")'
        ws.cell(r, 7).value = (
            f'=COUNTIFS(\'Bid Log\'!V:V,{oq},\'Bid Log\'!N:N,"Bid Submitted")'
            f'+COUNTIFS(\'Bid Log\'!V:V,{oq},\'Bid Log\'!N:N,"Budget Submitted")'
        )
        for ci in [3,4,5,6,7]:
            cell = ws.cell(r, ci)
            cell.font = F(9); cell.fill = BG(bg)
            cell.alignment = AL("left" if ci==3 else "center"); cell.border = BD()
        r += 1

# ══════════════════════════════════════════════════════════════════════════════
# DATA QUALITY TAB
# ══════════════════════════════════════════════════════════════════════════════
def build_dataquality(ws):
    ws.sheet_properties.tabColor = "FFC000"
    ws.sheet_view.showGridLines = False

    for i, w in enumerate([2,30,22,16,2], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    title_bar(ws, 1, "B", "D", "ASM — DATA QUALITY REPORT", bg=DKYEL, fg=WHITE, size=13)

    ws.row_dimensions[2].height = 18
    ws.merge_cells("B2:D2")
    c = ws["B2"]
    c.value = "Fix all flagged items before the Friday estimating meeting. Pull the Bid Log and filter Data Quality Flag ≠ OK."
    c.font = F(10, False, DGRAY, True); c.alignment = AL("left","center")

    checks = [
        ("REQUIRED FIELDS",            NAVY,  [
            ("Bids missing Project Name",
             '=COUNTIF(\'Bid Log\'!B:B,"")'),
            ("Bids missing General Contractor",
             '=COUNTIFS(\'Bid Log\'!C:C,"",\'Bid Log\'!B:B,"<>")'),
            ("Bids missing Estimator",
             '=COUNTIFS(\'Bid Log\'!I:I,"",\'Bid Log\'!B:B,"<>")'),
            ("Bids missing Bid Due Date",
             '=COUNTIFS(\'Bid Log\'!K:K,"",\'Bid Log\'!B:B,"<>")'),
            ("Bids missing ASM Bid Amount",
             '=COUNTIFS(\'Bid Log\'!M:M,"",\'Bid Log\'!B:B,"<>")'),
        ]),
        ("AWARDED BIDS",               DKGRN, [
            ("Awarded bids with no Award Amount",
             '=COUNTIFS(\'Bid Log\'!N:N,"Awarded",\'Bid Log\'!O:O,"")'),
        ]),
        ("LOST BIDS",                  DKRED, [
            ("Lost bids with no Lost Reason",
             '=COUNTIFS(\'Bid Log\'!N:N,"Lost",\'Bid Log\'!R:R,"")'),
        ]),
        ("FOLLOW-UP COMPLETENESS",     ORANGE, [
            ("Pending bids with no Follow-Up Date",
             '=COUNTIFS(\'Bid Log\'!AB:AB,"MISSING")'),
            ("Pending bids with no Follow-Up Owner",
             '=COUNTIFS(\'Bid Log\'!V:V,"",\'Bid Log\'!N:N,"Bid Submitted")+COUNTIFS(\'Bid Log\'!V:V,"",\'Bid Log\'!N:N,"Budget Submitted")'),
            ("Follow-ups past due",
             '=COUNTIF(\'Bid Log\'!AB:AB,"OVERDUE")'),
        ]),
        ("BID DUE DATE ALERTS",        BLUE,  [
            ("Active bids with Bid Due Date already passed",
             '=COUNTIFS(\'Bid Log\'!K:K,"<"&TODAY(),\'Bid Log\'!N:N,"In Progress")'),
            ("Pending bids older than 60 days with no activity",
             '=COUNTIFS(\'Bid Log\'!AA:AA,">60",\'Bid Log\'!AA:AA,"<>")'),
        ]),
        ("DATA QUALITY SUMMARY",       DGRAY, [
            ("Total rows with ANY data quality issue",
             '=COUNTIF(\'Bid Log\'!AC:AC,"<>OK")-1'),  # -1 for header
        ]),
    ]

    r = 3
    for section_name, color, items in checks:
        ws.row_dimensions[r].height = 8; r += 1
        section_bar(ws, r, "B", "D", f"  {section_name}", bg=color); r += 1
        ws.row_dimensions[r].height = 18
        hcell(ws.cell(r, 2), "Check", bg=BLUE, size=9)
        hcell(ws.cell(r, 3), "Count", bg=BLUE, size=9)
        hcell(ws.cell(r, 4), "Status", bg=BLUE, size=9)
        r += 1
        for label, formula in items:
            ws.row_dimensions[r].height = 18
            bg = LGRAY if r % 2 == 0 else WHITE
            vcell(ws.cell(r, 2), label, bg=bg, size=9, wrap=True)
            ws.cell(r, 3).value = formula
            ws.cell(r, 3).font  = F(11, True)
            ws.cell(r, 3).fill  = BG(bg)
            ws.cell(r, 3).alignment = AL("center")
            ws.cell(r, 3).border = BD()
            ws.cell(r, 3).number_format = "0"
            ws.cell(r, 4).value = f'=IF(C{r}=0,"✓ OK","⚠ {chr(38)} FIX NEEDED")'
            ws.cell(r, 4).font  = F(9, True)
            ws.cell(r, 4).fill  = BG(bg)
            ws.cell(r, 4).alignment = AL("center")
            ws.cell(r, 4).border = BD()
            # Conditional formatting on status cell
            ws.conditional_formatting.add(f"D{r}:D{r}", FormulaRule(
                formula=[f'C{r}>0'],
                dxf=DifferentialStyle(
                    font=Font(name="Arial", size=9, bold=True, color=DKRED),
                    fill=PatternFill("solid", fgColor=LTRED)),
                stopIfTrue=False))
            ws.conditional_formatting.add(f"D{r}:D{r}", FormulaRule(
                formula=[f'C{r}=0'],
                dxf=DifferentialStyle(
                    font=Font(name="Arial", size=9, bold=True, color=DKGRN),
                    fill=PatternFill("solid", fgColor=LTGRN)),
                stopIfTrue=False))
            r += 1

# ══════════════════════════════════════════════════════════════════════════════
# INSTRUCTIONS TAB
# ══════════════════════════════════════════════════════════════════════════════
def build_instructions(ws):
    ws.sheet_properties.tabColor = "4472C4"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 70

    title_bar(ws, 1, "B", "C",
        "ASM PRECONSTRUCTION PIPELINE TRACKER V9  —  INSTRUCTIONS & SOP",
        bg=NAVY, size=13)

    r = 2

    def inst_section(title, bg=NAVY):
        nonlocal r
        ws.row_dimensions[r].height = 8; r += 1
        section_bar(ws, r, "B", "C", f"  {title}", bg=bg, height=24); r += 1

    def inst_row(label, content, bg=None):
        nonlocal r
        ws.row_dimensions[r].height = 42
        bg_use = bg if bg else (LGRAY if r % 2 == 0 else WHITE)
        c1 = ws.cell(r, 2)
        c1.value = label; c1.font = F(10, True)
        c1.fill = BG(bg_use); c1.alignment = AL("left","top", True)
        c1.border = BD()
        c2 = ws.cell(r, 3)
        c2.value = content; c2.font = F(10)
        c2.fill = BG(bg_use); c2.alignment = AL("left","top", True)
        c2.border = BD()
        r += 1

    inst_section("WHO UPDATES THIS TRACKER")
    inst_row("Primary Owner",
        "The assigned estimator updates their own bids. The Preconstruction Manager or VP of Preconstruction reviews and spot-checks weekly.")
    inst_row("Backup / Admin",
        "If an estimator is unavailable, the Preconstruction Manager or office admin updates the tracker based on team notes.")

    inst_section("WHEN TO UPDATE")
    inst_row("When a bid is received",
        "Add a new row to the Bid Log immediately. Fill in: Project Name, GC, Location, Market Segment, Scope Type, Bid Stage, Estimator, Bid Received Date, Bid Due Date, Status = In Progress.")
    inst_row("When a bid is submitted",
        "Update: Submitted Date, ASM Bid Amount, Status = Bid Submitted (or Budget Submitted / ROM Submitted). Set Next Follow-Up Date (typically 5–7 business days after submission). Assign Follow-Up Owner.")
    inst_row("When following up",
        "Update: Last Contact Date, Notes (brief summary of conversation). Reset Next Follow-Up Date to next expected contact.")
    inst_row("When awarded",
        "Update: Status = Awarded, ASM Awarded Amount, Last Contact Date. Clear follow-up fields.")
    inst_row("When lost",
        "Update: Status = Lost, Winning Competitor, Competitor Amount (if known), Lost Reason. Add a note with any intel gathered.")
    inst_row("Before the Friday Meeting",
        "All estimators must update their bids by Thursday EOD. Check the Data Quality tab — all flags should read OK before the meeting.")

    inst_section("STATUS DEFINITIONS")
    statuses = [
        ("In Progress",       "Bid drawings have been received. Takeoff and pricing are underway. Bid has not yet been submitted."),
        ("ROM Submitted",     "Rough Order of Magnitude pricing has been provided to the GC or owner. Not a formal bid."),
        ("Budget Submitted",  "Preliminary budget pricing has been submitted. Often used in early design phases (SD/DD)."),
        ("Bid Submitted",     "A formal bid has been submitted. Awaiting GC or owner decision. Follow-up required."),
        ("Awarded",           "ASM has been selected for the project. Contract or LOI received."),
        ("Lost",              "GC or owner awarded the work to a competitor. Fill in Losing Competitor, Amount, and Lost Reason."),
        ("No Bid",            "ASM chose not to submit a bid. Reason should be noted in the Notes column."),
        ("On Hold",           "Project is paused, on hold, or decision has been delayed. Monitor for reactivation."),
    ]
    for status, definition in statuses:
        inst_row(status, definition)

    inst_section("GO / NO-GO SCORING GUIDE")
    inst_row("Purpose",
        "Use the Go-No-Go tab to score each bid before submitting. This helps leadership decide whether ASM should invest estimating time on a project.")
    inst_row("How to score",
        "Score each of the 8 categories from 1 (very low) to 5 (very high). The spreadsheet calculates a total out of 40 and shows a recommendation automatically.")
    inst_row("Recommendations",
        "GO = 70% or higher (score 28+).  REVIEW = 50-69% (score 20-27) — bring to leadership for a decision.  NO-GO = below 50% (score under 20) — recommend passing unless strategic reason exists.")
    inst_row("Override policy",
        "A leadership override can change a No-Go to a Go for strategic reasons. Document the reason in the Notes column of the Bid Log.")

    inst_section("FRIDAY MEETING AGENDA GUIDE")
    inst_row("Before the meeting (Thursday EOD)",
        "1. Update all bid statuses.\n2. Set next follow-up dates for all pending bids.\n3. Check Data Quality tab — all rows should show OK.\n4. Score any new bids in the Go-No-Go tab.")
    inst_row("At the meeting — review these tabs",
        "1. Dashboard — overall pipeline health and win rate.\n2. Friday Meeting tab — bids due, follow-ups, expected awards.\n3. Go-No-Go tab — any No-Go or Review items.\n4. Follow-Up Control tab — overdue or missing follow-ups.")
    inst_row("Key questions to answer",
        "1. Are we bidding enough work to feed operations?\n2. Do any estimators need workload relief?\n3. Which pending bids are most likely to convert this month?\n4. Are there any stale bids (60+ days) we should close out?\n5. What did we learn from recent losses?")

    inst_section("TIPS FOR KEEPING THE TRACKER CLEAN")
    inst_row("Use dropdowns",
        "Always select from the dropdown lists. Do not type statuses, market segments, or estimator names manually — this will break filters and dashboard formulas.")
    inst_row("One row per bid",
        "Each bid gets one row. If a project re-bids or has a scope revision, add a new row with a new Bid ID and note the relationship in the Notes column.")
    inst_row("Don't delete rows",
        "Lost and No Bid rows should stay in the tracker — they are needed for win rate calculations and historical analysis.")
    inst_row("Probability guidelines",
        "0% = No Bid or Lost.  25% = Long shot / weak relationship.  50% = Competitive with no clear advantage.  75% = Strong position / good GC relationship.  100% = Awarded.")

# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════
def main():
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    # Create sheets in display order
    ws_dash  = wb.create_sheet("Dashboard")
    ws_bid   = wb.create_sheet("Bid Log")
    ws_gng   = wb.create_sheet("Go-No-Go")
    ws_fri   = wb.create_sheet("Friday Meeting")
    ws_fu    = wb.create_sheet("Follow-Up Control")
    ws_dq    = wb.create_sheet("Data Quality")
    ws_set   = wb.create_sheet("Settings")
    ws_inst  = wb.create_sheet("Instructions")

    print("Building Settings...")
    build_settings(ws_set)
    print("Building Bid Log...")
    build_bid_log(ws_bid)
    print("Building Go-No-Go...")
    build_gonogo(ws_gng)
    print("Building Dashboard...")
    build_dashboard(ws_dash)
    print("Building Friday Meeting...")
    build_friday(ws_fri)
    print("Building Follow-Up Control...")
    build_followup(ws_fu)
    print("Building Data Quality...")
    build_dataquality(ws_dq)
    print("Building Instructions...")
    build_instructions(ws_inst)

    # Set Dashboard as active tab
    ws_dash.sheet_view.tabSelected = True
    wb.active = wb.index(ws_dash)

    wb.save(OUTPUT)
    print(f"Saved: {OUTPUT}")

if __name__ == "__main__":
    main()
